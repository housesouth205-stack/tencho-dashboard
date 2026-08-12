#!/usr/bin/env python3
"""パチスロ機種データベース ビルダー

data/machines.json を唯一の入力とし、以下を生成する。

  output/slot_machine_database.csv   機種マスター
  output/slot_machine_database.xlsx  4シート（機種マスター／出率ランキング／要確認／未取得機種）
  output/unacquired_machines.csv     未取得機種一覧
  output/needs_review.csv            要確認機種一覧
  output/quality_report.txt          データ品質チェック結果

方針：値の補完・推測は一切行わない。欠損は空欄のまま出力し、
異常値は自動修正せず「要確認」として列挙する。
"""

from __future__ import annotations

import json
import re
import sys
from collections import Counter
from pathlib import Path
from urllib.parse import urlparse

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data" / "machines.json"
OUT = ROOT / "output"

SETTING_COLS = [f"設定{i}出率" for i in range(1, 7)]

COLUMNS = [
    "機種名",
    "メーカー",
    "機種タイプ",
    "メディア区分",
    "ATタイプ",
    "ボーナスタイプ",
    "規則区分",
    "型式名",
    "コイン単価",
    "コイン単価条件",
    *SETTING_COLS,
    "出率条件",
    "出率出典URL",
    "コイン単価出典URL",
    "メーカー出典URL",
    "データ取得日",
    "現行判定",
    "信頼度",
    "備考",
    "別表記",
]

# ---------------------------------------------------------------- 正規化辞書

# メーカー名の表記ゆれ統一（左＝ゆれ、右＝正式表記）
MAKER_CANON = {
    "SANKYO": "SANKYO",
    "三共": "SANKYO",
    "サンキョー": "SANKYO",
    "Sammy": "サミー",
    "サミー": "サミー",
    "SAMMY": "サミー",
    "北電子": "北電子",
    "KITA DENSHI": "北電子",
    "山佐": "山佐",
    "YAMASA": "山佐",
    "ヤマサ": "山佐",
    "大都技研": "大都技研",
    "DAITO": "大都技研",
    "ビスティ": "ビスティ",
    "Bisty": "ビスティ",
    "パイオニア": "パイオニア",
    "PIONEER": "パイオニア",
    "藤商事": "藤商事",
    "FUJISHOJI": "藤商事",
    "ミズホ": "ミズホ",
    "メーシー": "メーシー",
    "オリンピア": "オリンピア",
    "エンターライズ": "エンターライズ",
    "不明": "不明",
}

# 機種タイプの表記ゆれ統一
TYPE_CANON = {
    "AT": "AT",
    "AT機": "AT",
    "ART": "ART",
    "ART機": "ART",
    "A+AT": "A+AT",
    "A＋AT": "A+AT",
    "AT+ART": "A+AT",
    "Aタイプ": "Aタイプ",
    "ノーマル": "Aタイプ",
    "ノーマルタイプ": "Aタイプ",
    "A": "Aタイプ",
    "BT": "BT",
    "ボーナスタイプ": "BT",
    "不明": "不明",
}

MEDIA_CANON = {
    "スマスロ": "スマスロ",
    "スマートパチスロ": "スマスロ",
    "メダル機": "メダル機",
    "メダル": "メダル機",
    "不明": "不明",
}

CONFIDENCE_ORDER = {"高": 0, "中": 1, "低": 2, "要確認": 3}

# 出率の常識的レンジ（これを外れたら異常値として要確認に回す）
RATE_MIN, RATE_MAX = 80.0, 130.0
# コイン単価の常識的レンジ（円）
COIN_MIN, COIN_MAX = 0.5, 10.0


def canon(value: str, table: dict[str, str], label: str, issues: list[str], machine: str) -> str:
    """表記ゆれを辞書で統一する。辞書に無い値はそのまま通し、品質チェックで報告。"""
    v = (value or "").strip()
    if not v:
        return ""
    if v in table:
        return table[v]
    issues.append(f"[表記ゆれ未登録] {machine}: {label}='{v}' は正規化辞書に未登録")
    return v


def join_urls(urls) -> str:
    if not urls:
        return ""
    return " | ".join(urls)


def valid_url(u: str) -> bool:
    try:
        p = urlparse(u)
        return p.scheme in ("http", "https") and bool(p.netloc)
    except ValueError:
        return False


def num_or_blank(v):
    """数値なら float、空文字/None/'不明' なら空文字を返す。推測補完はしない。"""
    if v is None or v == "" or v == "不明":
        return ""
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("%", "")
    try:
        return float(s)
    except ValueError:
        return ""


def build_master(raw: dict, issues: list[str]) -> pd.DataFrame:
    acquired_at = raw["_meta"]["取得日"]
    rows = []
    for m in raw["machines"]:
        name = m["機種名"]
        row = {
            "機種名": name,
            "メーカー": canon(m.get("メーカー", ""), MAKER_CANON, "メーカー", issues, name),
            "機種タイプ": canon(m.get("機種タイプ", ""), TYPE_CANON, "機種タイプ", issues, name),
            "メディア区分": canon(m.get("メディア区分", ""), MEDIA_CANON, "メディア区分", issues, name),
            "ATタイプ": m.get("ATタイプ", ""),
            "ボーナスタイプ": m.get("ボーナスタイプ", ""),
            "規則区分": m.get("規則区分", ""),
            "型式名": m.get("型式名", ""),
            "コイン単価": num_or_blank(m.get("コイン単価")),
            "コイン単価条件": m.get("コイン単価条件", ""),
            "出率条件": m.get("出率条件", ""),
            "出率出典URL": join_urls(m.get("出率出典URL")),
            "コイン単価出典URL": join_urls(m.get("コイン単価出典URL")),
            "メーカー出典URL": join_urls(m.get("メーカー出典URL")),
            "データ取得日": acquired_at,
            "現行判定": m.get("現行判定", ""),
            "信頼度": m.get("信頼度", ""),
            "備考": m.get("備考", ""),
            "別表記": " / ".join(m.get("別表記", [])),
        }
        for c in SETTING_COLS:
            row[c] = num_or_blank(m.get(c))
        rows.append(row)

    df = pd.DataFrame(rows, columns=COLUMNS)
    return df


# ---------------------------------------------------------------- 品質チェック


def quality_checks(df: pd.DataFrame, raw: dict) -> tuple[list[str], set[str]]:
    """品質チェックを実行。(問題行リスト, 要確認に回す機種名の集合) を返す。"""
    issues: list[str] = []
    flagged: set[str] = set()

    def flag(name: str, msg: str):
        issues.append(f"[{name}] {msg}")
        flagged.add(name)

    # 1. 機種名の重複（正規表記・別表記の両方で判定）
    dup = df["機種名"][df["機種名"].duplicated()].tolist()
    for d in dup:
        flag(d, "機種名が2行以上存在する（重複排除もれ）")

    alias_map: dict[str, str] = {}
    for _, r in df.iterrows():
        keys = [r["機種名"]] + [a for a in str(r["別表記"]).split(" / ") if a]
        for k in keys:
            norm = re.sub(r"[\sＬL　]|スマスロ|パチスロ", "", k)
            if not norm:
                continue
            if norm in alias_map and alias_map[norm] != r["機種名"]:
                flag(r["機種名"], f"別表記『{k}』が既存機種『{alias_map[norm]}』と衝突（同一機種の可能性）")
            alias_map.setdefault(norm, r["機種名"])

    for _, r in df.iterrows():
        name = r["機種名"]
        vals = {c: r[c] for c in SETTING_COLS}
        filled = {c: v for c, v in vals.items() if v != ""}

        # 2. 設定1〜6が全て入っているか
        if len(filled) == 0:
            flag(name, "設定1〜6の出率が1つも入っていない（データ欠損）")
        elif len(filled) < 6:
            missing = [c for c in SETTING_COLS if vals[c] == ""]
            flag(name, f"出率欠損: {', '.join(missing)}")

        # 3. 設定1より設定6が低い／単調性の破綻
        s1, s6 = vals["設定1出率"], vals["設定6出率"]
        if s1 != "" and s6 != "" and s6 < s1:
            flag(name, f"設定6({s6}%)が設定1({s1}%)より低い（異常）")
        ordered = [(c, v) for c, v in vals.items() if v != ""]
        for (ca, va), (cb, vb) in zip(ordered, ordered[1:]):
            if vb < va:
                flag(name, f"{ca}({va}%) > {cb}({vb}%) と逆転している（要確認）")

        # 4. 出率の異常値
        for c, v in filled.items():
            if not (RATE_MIN <= v <= RATE_MAX):
                flag(name, f"{c}={v}% が想定レンジ({RATE_MIN}〜{RATE_MAX}%)外")

        # 5. コイン単価の異常値／未取得
        coin = r["コイン単価"]
        if coin == "":
            flag(name, "コイン単価未確認")
        elif not (COIN_MIN <= coin <= COIN_MAX):
            flag(name, f"コイン単価={coin}円 が想定レンジ({COIN_MIN}〜{COIN_MAX}円)外")

        # 6. URLの存在と妥当性
        for col in ("出率出典URL", "コイン単価出典URL", "メーカー出典URL"):
            urls = [u for u in str(r[col]).split(" | ") if u]
            if col == "出率出典URL" and not urls and len(filled) > 0:
                flag(name, "出率の値があるのに出典URLが無い")
            if col == "メーカー出典URL" and not urls:
                flag(name, "メーカー出典URLが無い")
            for u in urls:
                if not valid_url(u):
                    flag(name, f"{col} に不正なURL: {u}")

        # 7. 出率の情報源が1つしかない
        rate_urls = [u for u in str(r["出率出典URL"]).split(" | ") if u]
        if len(filled) > 0 and len(rate_urls) < 2:
            flag(name, "出率の情報源が1つしかない（照合未達）")

        # 8. 出率条件・現行判定・メーカーの不明
        if len(filled) > 0 and (not r["出率条件"] or "不明" in str(r["出率条件"])):
            flag(name, "出率条件が不明")
        if "要確認" in str(r["現行判定"]):
            flag(name, "現行機かどうか不明（現行判定要確認）")
        if r["メーカー"] in ("", "不明"):
            flag(name, "メーカー不明")
        if r["機種タイプ"] in ("", "不明"):
            flag(name, "機種タイプ不明")

    return issues, flagged


# ---------------------------------------------------------------- シート生成


def build_rankings(df: pd.DataFrame) -> pd.DataFrame:
    """出率ランキングシート。値が無い機種はランキングから除外（推測補完しない）。"""
    blocks = []

    def add(title: str, series_col: str, frame: pd.DataFrame, value_name: str):
        sub = frame[frame[series_col] != ""].copy()
        if sub.empty:
            blocks.append(pd.DataFrame({"ランキング": [title], "順位": ["該当データなし"],
                                        "機種名": [""], "メーカー": [""], value_name: [""]}))
            return
        sub = sub.sort_values(series_col, ascending=False)
        sub.insert(0, "順位", range(1, len(sub) + 1))
        out = sub[["順位", "機種名", "メーカー", series_col]].rename(columns={series_col: value_name})
        out.insert(0, "ランキング", "")
        out.loc[out.index[0], "ランキング"] = title
        blocks.append(out)
        blocks.append(pd.DataFrame({"ランキング": [""], "順位": [""], "機種名": [""],
                                    "メーカー": [""], value_name: [""]}))

    frames = []
    for setting in ("設定1出率", "設定4出率", "設定5出率", "設定6出率"):
        tmp = df.copy()
        add(f"{setting}ランキング", setting, tmp, "値(%)")
        frames.append(None)

    # 設定1→設定6の出率差
    diff = df.copy()
    diff["設定1-6差"] = [
        (r["設定6出率"] - r["設定1出率"]) if (r["設定1出率"] != "" and r["設定6出率"] != "") else ""
        for _, r in diff.iterrows()
    ]
    add("設定1→設定6 出率差ランキング", "設定1-6差", diff, "値(%)")
    add("コイン単価ランキング", "コイン単価", df.copy(), "値(%)")

    result = pd.concat(blocks, ignore_index=True)
    return result.rename(columns={"値(%)": "値"})


def build_needs_review(df: pd.DataFrame, issues: list[str], flagged: set[str]) -> pd.DataFrame:
    by_machine: dict[str, list[str]] = {}
    for msg in issues:
        m = re.match(r"\[(.+?)\]\s*(.*)", msg)
        if not m:
            continue
        by_machine.setdefault(m.group(1), []).append(m.group(2))

    rows = []
    for _, r in df.iterrows():
        name = r["機種名"]
        if name not in by_machine:
            continue
        rows.append({
            "機種名": name,
            "メーカー": r["メーカー"],
            "信頼度": r["信頼度"],
            "要確認項目数": len(by_machine[name]),
            "要確認内容": " ／ ".join(by_machine[name]),
            "出率出典URL": r["出率出典URL"],
            "データ取得日": r["データ取得日"],
        })
    out = pd.DataFrame(rows)
    if not out.empty:
        out = out.sort_values("要確認項目数", ascending=False).reset_index(drop=True)
    return out


def build_unacquired(raw: dict) -> pd.DataFrame:
    rows = []
    for m in raw["machines"]:
        if not m.get("未取得理由"):
            continue
        rows.append({
            "機種名": m["機種名"],
            "メーカー": m.get("メーカー", ""),
            "取得できなかった理由": m["未取得理由"],
            "試した情報源": " / ".join(m.get("試した情報源", [])),
            "URL": join_urls(m.get("メーカー出典URL")),
            "再調査が必要か": m.get("再調査要否", "要"),
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------- Excel 出力


HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def style_sheet(ws, df: pd.DataFrame):
    if df.empty:
        return
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, col in enumerate(df.columns, start=1):
        longest = max([len(str(col))] + [len(str(v)) for v in df[col].head(200)])
        ws.column_dimensions[get_column_letter(i)].width = min(max(longest + 2, 10), 60)


def write_excel(path: Path, master, rankings, review, unacquired):
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        master.to_excel(xw, sheet_name="機種マスター", index=False)
        rankings.to_excel(xw, sheet_name="出率ランキング", index=False)
        (review if not review.empty else pd.DataFrame({"機種名": ["該当なし"]})).to_excel(
            xw, sheet_name="要確認", index=False)
        (unacquired if not unacquired.empty else pd.DataFrame({"機種名": ["該当なし"]})).to_excel(
            xw, sheet_name="未取得機種", index=False)

        style_sheet(xw.sheets["機種マスター"], master)
        style_sheet(xw.sheets["出率ランキング"], rankings)
        style_sheet(xw.sheets["要確認"], review)
        style_sheet(xw.sheets["未取得機種"], unacquired)


# ---------------------------------------------------------------- レポート


def summary(df: pd.DataFrame, review: pd.DataFrame, unacquired: pd.DataFrame) -> list[str]:
    total = len(df)
    complete = sum(1 for _, r in df.iterrows() if all(r[c] != "" for c in SETTING_COLS))
    partial = sum(1 for _, r in df.iterrows()
                  if any(r[c] != "" for c in SETTING_COLS) and any(r[c] == "" for c in SETTING_COLS))
    none_got = sum(1 for _, r in df.iterrows() if all(r[c] == "" for c in SETTING_COLS))
    coin_got = sum(1 for _, r in df.iterrows() if r["コイン単価"] != "")

    domains = Counter()
    for _, r in df.iterrows():
        for col in ("出率出典URL", "コイン単価出典URL", "メーカー出典URL"):
            for u in str(r[col]).split(" | "):
                if u:
                    domains[urlparse(u).netloc] += 1

    cross = sum(1 for _, r in df.iterrows()
                if len([u for u in str(r["出率出典URL"]).split(" | ") if u]) >= 2)

    lines = [
        "===== 最終報告 =====",
        f"調査対象機種数        : {total}",
        f"完全取得（設定1〜6全）: {complete}",
        f"一部欠損機種数        : {partial}",
        f"未取得機種数（出率0件）: {none_got}",
        f"要確認機種数          : {len(review)}",
        f"コイン単価取得率      : {coin_got}/{total} ({coin_got / total * 100:.1f}%)" if total else "",
        f"出率の照合率(2URL以上): {cross}/{total} ({cross / total * 100:.1f}%)" if total else "",
        "",
        "--- 情報源別 取得件数（ドメイン単位） ---",
    ]
    for d, c in domains.most_common():
        lines.append(f"  {d}: {c}")
    if domains:
        lines.append(f"最も多く使用した情報源: {domains.most_common(1)[0][0]}")

    lines.append("")
    lines.append("--- 信頼度分布 ---")
    for k, c in df["信頼度"].value_counts().items():
        lines.append(f"  {k}: {c}")
    return [line for line in lines if line != ""] + [""]


def main() -> int:
    raw = json.loads(DATA.read_text(encoding="utf-8"))
    OUT.mkdir(exist_ok=True)

    canon_issues: list[str] = []
    master = build_master(raw, canon_issues)

    issues, flagged = quality_checks(master, raw)
    rankings = build_rankings(master)
    review = build_needs_review(master, issues, flagged)
    unacquired = build_unacquired(raw)

    master.to_csv(OUT / "slot_machine_database.csv", index=False, encoding="utf-8-sig")
    review.to_csv(OUT / "needs_review.csv", index=False, encoding="utf-8-sig")
    unacquired.to_csv(OUT / "unacquired_machines.csv", index=False, encoding="utf-8-sig")
    write_excel(OUT / "slot_machine_database.xlsx", master, rankings, review, unacquired)

    report = []
    report.append("===== データ品質チェック =====")
    report.append(f"検出件数: {len(issues) + len(canon_issues)}")
    report.append("")
    report.append("--- 表記ゆれチェック ---")
    report.extend(canon_issues or ["  問題なし（全て正規化辞書に登録済み）"])
    report.append("")
    report.append("--- 機種別チェック ---")
    report.extend(f"  {i}" for i in issues)
    report.append("")
    report.append("※ 検出された異常値は自動修正せず、全て『要確認』シートに転記済み。")
    report.append("")
    report.extend(summary(master, review, unacquired))

    text = "\n".join(report)
    (OUT / "quality_report.txt").write_text(text, encoding="utf-8")
    print(text)
    return 0


if __name__ == "__main__":
    sys.exit(main())
